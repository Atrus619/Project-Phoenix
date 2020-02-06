import pandas as pd
import nltk
from openpyxl import load_workbook
from prefect import task, utilities
from config import Config as cfg


@task
def utterances_to_tokens(path, reuse_existing=True, remove_caps=True):
    """
    Adds a sheet to excel file containing utterances for training intent classifier for the purpose of creating labels for NER
    :param path: String, Path to excel sheet
    :param reuse_existing: Boolean, Whether to reuse existing annotations if present
    :param remove_caps: Boolean, Whether to convert text to lower case
    :param logger: Optional logger
    :return: Modifies excel sheet in place. Returns True if successful.
    """
    logger = utilities.logging.get_logger(cfg.chatbot_training_log_name)
    logger.info(f'----- Processing input excel sheet from {path} to convert utterances to tokens -----')

    book = load_workbook(path)
    assert 'TrainingExamples' in book.sheetnames

    df = pd.read_excel(path, sheet_name='TrainingExamples')
    new_df = pd.DataFrame(columns=['OG_Text', 'Token', 'Label'])

    # If Tokens sheet already exists, reuse what has already been filled out
    reused_tokens = None
    if 'Tokens' in book.sheetnames and reuse_existing:
        old_tokens = pd.read_excel(path, sheet_name='Tokens')
        logger.info(f'{old_tokens.shape[0]} tokens found in existing worksheet. Transferring over prior entries that remain in TrainingExamples sheet.')

        # Only copy over tokens that exist in the current list of training examples
        previously_tokenized_phrases = set(df.TrainingExample)
        reused_tokens = old_tokens[old_tokens.OG_Text.isin(previously_tokenized_phrases)]
        logger.info(f'{reused_tokens.shape[0]} / {old_tokens.shape[0]} previously labeled tokens retained.')

        new_df = new_df.append(reused_tokens)
        new_df = new_df.fillna('')
        existing_phrases = set(new_df.OG_Text)
    else:
        existing_phrases = {}

    for i, example in df.iterrows():
        if example[1] not in existing_phrases:
            if remove_caps:
                text = example[1].lower()
            else:
                text = example[1]
            tokens = nltk.word_tokenize(text)
            for token in tokens:
                new_df = new_df.append(pd.DataFrame({
                    'OG_Text': [example[1]],
                    'Token': [token],
                    'Label': ['O']
                }))
            # Stanford NER requires separating each phrase with a blank space row
            new_df = new_df.append(pd.DataFrame({
                'OG_Text': [example[1]],
                'Token': [''],
                'Label': ['']
            }))

    new_entries = new_df.shape[0] - (reused_tokens.shape[0] if reused_tokens is not None else 0)
    logger.info(f'{new_entries} new token entries added to Tokens sheet.')

    # Overwrite file with new sheet Tokens
    if 'Tokens' in book.sheetnames:
        del book['Tokens']

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = book
        new_df.to_excel(writer, sheet_name='Tokens', index=False)
        writer.save()

    logger.info(f'----- Conversion of utterances to tokens successful. Please refer to updated file at {path}. -----')

    return
